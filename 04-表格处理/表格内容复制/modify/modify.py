import xlsxwriter
from openpyxl import load_workbook

# 加载现有的Excel文件
input_file = 'input.xlsx'
wb = load_workbook(input_file)

# 创建一个新的Excel文件
output_file = 'output.xlsx'
workbook = xlsxwriter.Workbook(output_file)

# 遍历所有工作表
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    worksheet = workbook.add_worksheet(sheet_name)  # 创建与原工作表同名的新工作表

    # 复制原始工作表的数据
    for row in ws.iter_rows():
        for cell in row:
            # 写入数据
            worksheet.write(cell.row - 1, cell.column - 1, cell.value)

    # 读取A6单元格的内容
    cell_value = ws['A6'].value
    if cell_value:
        parts = cell_value.split('：', 1)  # 切割内容为两部分
        if len(parts) < 2:
            parts.append('：')  # 如果没有找到分隔符，则默认第二部分为空
        text_part1, text_part2 = parts

        # 打印分割后的内容（用于调试）
        print(f"Sheet: {sheet_name}")
        print("Text Part 1:", text_part1)
        print("Text Part 2:", text_part2)

        # 定义字体样式
        font1_format = workbook.add_format({'font_name': '宋体', 'bold': True, 'size': 12})
        font2_format = workbook.add_format({'font_name': '微软雅黑', 'bold': True, 'size': 12})

        # 将带有不同字体样式的内容写入到A6单元格
        worksheet.write_rich_string(
            5, 0,  # A6对应的行列索引
            text_part1, font1_format, '：',  # 添加分隔符
            text_part2, font2_format
        )

# 关闭并保存工作簿
workbook.close()
