from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 加载现有的Excel文件
input_file = 'output.xlsx'
wb = load_workbook(input_file)

# 定义格式
bold_font = Font(bold=True, name='微软雅黑', size=12)
song_font = Font(name='宋体', size=12)
alignment_center = Alignment(horizontal='center', vertical='center')
alignment_left_center = Alignment(horizontal='left', vertical='center')
alignment_left_top = Alignment(horizontal='left', vertical='top')
thin_border = Border(
    top=Side(border_style="thin", color="000000"),
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# 遍历所有工作表
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # 合并单元格
    merge_cells = [
        'A3:B3', 'C3:E3', 'A4:B4', 'C4:D4', 'A5:B5', 'C5:D5',
        'A6:E6', 'C7:E7', 'C8:E8', 'A9:B9', 'C9:E9', 'A10:B10',
        'C10:E10', 'A11:E11', 'A12:E13'
    ]
    for merge_range in merge_cells:
        ws.merge_cells(merge_range)

    # 设置列宽和行高
    for col in range(1, 6):
        ws.column_dimensions[chr(col + 64)].width = 24.89
    for row in range(1, 13):
        ws.row_dimensions[row].height = 31.95

    # # 设置所有单元格的字体和对齐方式
    # for row in range(1, 14):
    #     for col in range(1, 6):
    #         # 跳过A6单元格
    #         if row == 6:
    #             continue
    #         cell = ws.cell(row=row, column=col)
    #         cell.font = song_font
    #         cell.alignment = alignment_center

    # # 特殊处理对齐方式
    # ws['A6'].alignment = alignment_left_center
    # ws['A11'].alignment = alignment_left_center
    # ws['A12'].alignment = alignment_left_top
    #
    # # 添加边框
    # for row in range(1, 14):
    #     for col in range(1, 6):
    #         cell = ws.cell(row=row, column=col)
    #         cell.border = thin_border
    #
    # # 设置加粗字体
    # bold_cells = ['A1', 'B1', 'C1', 'D1', 'E1', 'A3', 'A4', 'C4', 'E4', 'A7', 'B7', 'C7', 'A9', 'C9', 'A11',
    #               'A12']
    # for cell_address in bold_cells:
    #     ws[cell_address].font = bold_font
    #
    # # 设置非加粗的字体为宋体
    # song_cells = ['A2', 'B2', 'C2', 'D2', 'E2', 'C3', 'C5', 'C8']
    # for cell_address in song_cells:
    #     ws[cell_address].font = song_font

# 保存修改后的文件
output_file = 'output.xlsx'
wb.save(output_file)
