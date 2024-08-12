import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

from test2 import parse_excel


def create_sheet(num, longitude, latitude, town_name, ground_type, _2_soil_type, soil_species_names,
                 prediction_of_soil_species, change_pattern):
    # 写入子表DQ-01
    worksheet = workbook.create_sheet(title=num)

    # 填写表头
    worksheet['A1'] = '编号'
    worksheet['B1'] = '经度/(°E)'
    worksheet['C1'] = '纬度/(°N)'
    worksheet['D1'] = '村镇名'
    worksheet['E1'] = '地类'

    # 填写数据
    worksheet['A2'] = num
    worksheet['B2'] = float(longitude)
    worksheet['C2'] = float(latitude)
    worksheet['D2'] = town_name
    worksheet['E2'] = ground_type

    worksheet['A3'] = '二普土壤类型'
    worksheet['C3'] = _2_soil_type

    soil_species_names = _2_soil_type.split('-')[-1]

    worksheet['A4'] = '二普土种名'
    worksheet['C4'] = '省级修正土种名'
    worksheet['E4'] = '图斑面积/hm²'
    worksheet['C5'] = soil_species_names

    worksheet['A6'] = f'室内初步预测土种：{prediction_of_soil_species}'

    worksheet['A7'] = '土钻点经度/(°E)'
    worksheet['B7'] = '土钻点纬度/(°N)'
    worksheet['C7'] = '实际地类变化模式'
    worksheet['C8'] = change_pattern

    worksheet['A9'] = '调查日期'
    worksheet['C9'] = '校核专家'

    # worksheet['A10'] = ''
    worksheet['A11'] = '野外土种校核结果：'
    worksheet['A12'] = '备注：'

    # 合并单元格
    worksheet.merge_cells('A3:B3')
    worksheet.merge_cells('C3:E3')
    worksheet.merge_cells('A4:B4')
    worksheet.merge_cells('C4:D4')
    worksheet.merge_cells('A5:B5')
    worksheet.merge_cells('C5:D5')
    worksheet.merge_cells('A6:E6')
    worksheet.merge_cells('C7:E7')
    worksheet.merge_cells('C8:E8')
    worksheet.merge_cells('A9:B9')
    worksheet.merge_cells('C9:E9')
    worksheet.merge_cells('A10:B10')
    worksheet.merge_cells('C10:E10')
    worksheet.merge_cells('A11:E11')
    worksheet.merge_cells('A12:E13')

    # 设置列宽和行高
    for col in range(1, 6):
        worksheet.column_dimensions[chr(col + 64)].width = 24.89
    for row in range(1, 13):
        worksheet.row_dimensions[row].height = 31.95

    # 设置所有单元格的字体为微软雅黑,内容居中,12号字体
    for row in range(1, 14):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            cell.font = Font(name='微软雅黑', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 特殊处理A6的居中为左对齐
    worksheet['A6'].alignment = Alignment(horizontal='left', vertical='center')
    worksheet['A11'].alignment = Alignment(horizontal='left', vertical='center')
    # 特殊处理A12的居中为左对齐、上对齐
    worksheet['A12'].alignment = Alignment(horizontal='left', vertical='top')

    # 添加边框
    thin = Side(border_style="thin", color="000000")
    for row in range(1, 14):
        for col in range(1, 6):
            cell = worksheet.cell(row=row, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 设置A1单元格加粗
    bold_font = Font(bold=True, name='微软雅黑', size=12)
    worksheet['A1'].font = bold_font
    worksheet['B1'].font = bold_font
    worksheet['C1'].font = bold_font
    worksheet['D1'].font = bold_font
    worksheet['E1'].font = bold_font
    worksheet['A3'].font = bold_font
    worksheet['A4'].font = bold_font
    worksheet['C4'].font = bold_font
    worksheet['E4'].font = bold_font
    worksheet['A6'].font = bold_font
    worksheet['A7'].font = bold_font
    worksheet['B7'].font = bold_font
    worksheet['C7'].font = bold_font
    worksheet['A9'].font = bold_font
    worksheet['C9'].font = bold_font
    worksheet['A11'].font = bold_font
    worksheet['A12'].font = bold_font

    # 设置非加粗的字体为宋体
    song_font = Font(name='宋体', size=12)
    worksheet['A2'].font = song_font
    worksheet['B2'].font = song_font
    worksheet['C2'].font = song_font
    worksheet['D2'].font = song_font
    worksheet['E2'].font = song_font
    worksheet['C3'].font = song_font
    worksheet['C5'].font = song_font
    worksheet['A6'].font = song_font
    worksheet['C8'].font = song_font


if __name__ == '__main__':
    file_path = r"野外校核点属性表.xls"
    save_path = r"result.xlsx"

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    value_lst = parse_excel(file_path)
    for i, value in enumerate(value_lst):
        print(value)
        create_sheet(*value)

    # 保存工作簿
    workbook.save(save_path)
